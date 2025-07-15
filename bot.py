
from telegram.constants import ParseMode
from telegram.helpers import escape_markdown
from telegram.constants import ParseMode
from html import escape 
import admin
from admin import ADMIN_STATES, ADMIN_IDS, admin_panel, admin_callback_handler, admin_handle_text_message
    

import asyncio
import logging
import os
import io
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
from telegram.constants import ParseMode
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Logging sozlamalari
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Konfiguratsiya
BOT_TOKEN = "8024757573:AAHCh_-DlS823fwkttv_lRVGpCwpK-RsScY"
DB_CONFIG = {
    'host': 'localhost',
    'database': 'dorixona_bot',
    'user': 'postgres',
    'password': 'auktambek012',
    'port': 5432
}
GROUP_CHAT_ID = -1002163331333  # Buyurtmalar yuboriladigan guruh ID




# Ma'lumotlar bazasi bilan ishlash klassi
class DatabaseManager:
    def __init__(self, config):
        self.config = config



    def add_pharmacy(self, data: Dict) -> bool:
            """Yangi dorixona qo'shish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO dorixonalar (dorixona_nomi, manzil, inn, dorixona_egasi, kontrakt, dagovor, rs, mfo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (data['dorixona_nomi'], data['manzil'], data['inn'], data['dorixona_egasi'],
                             data['kontrakt'], data['dagovor'], data['rs'], data['mfo'])
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error adding pharmacy: {e}")
                return False

    def update_pharmacy(self, pharmacy_id: int, field_name: str, value: str) -> bool:
            """Dorixona ma'lumotini tahrirlash"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE dorixonalar SET {field_name} = %s WHERE id = %s",
                            (value, pharmacy_id)
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error updating pharmacy {field_name}: {e}")
                return False

    def delete_pharmacy(self, pharmacy_id: int) -> bool:
            """Dorixonani o'chirish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM dorixonalar WHERE id = %s", (pharmacy_id,))
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error deleting pharmacy: {e}")
                return False

    def get_all_employees(self) -> List[Dict]:
            """Barcha xodimlarni olish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("SELECT * FROM xodimlar")
                        return [dict(row) for row in cur.fetchall()]
            except Exception as e:
                logger.error(f"Database error in get_all_employees: {e}")
                return []

    def add_employee(self, data: Dict) -> bool:
            """Yangi xodim qo'shish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO xodimlar (ism_familiya, telefon_raqam, id)
                            VALUES (%s, %s, %s)
                            """,
                            (data['ism_familiya'], data['telefon_raqam'], data['id'])
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error adding employee: {e}")
                return False

    def update_employee(self, employee_id: int, field_name: str, value: str) -> bool:
            """Xodim ma'lumotini tahrirlash"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE xodimlar SET {field_name} = %s WHERE id = %s",
                            (value, employee_id)
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error updating employee {field_name}: {e}")
                return False

    def delete_employee(self, employee_id: int) -> bool:
            """Xodimni o'chirish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM xodimlar WHERE id = %s", (employee_id,))
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error deleting employee: {e}")
                return False

    def add_medicine(self, data: Dict) -> bool:
            """Yangi dori qo'shish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO dorilar (dori_nomi, narxi, ikpu, image_file_id, info)
                            VALUES (%s, %s, %s, %s, %s)
                            """,
                            (data['dori_nomi'], data['narxi'], data['ikpu'], data['image_file_id'], data['info'])
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error adding medicine: {e}")
                return False

    def update_medicine(self, medicine_id: int, field_name: str, value) -> bool:
            """Dori ma'lumotini tahrirlash"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE dorilar SET {field_name} = %s WHERE id = %s",
                            (value, medicine_id)
                        )
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error updating medicine {field_name}: {e}")
                return False

    def delete_medicine(self, medicine_id: int) -> bool:
            """Dorini o'chirish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM dorilar WHERE id = %s", (medicine_id,))
                        conn.commit()
                return True
            except Exception as e:
                logger.error(f"Error deleting medicine: {e}")
                return False
    


    









    def get_all_pharmacies(self):
            """Barcha dorixonalarni olish"""
            try:
                with self.get_connection() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("SELECT * FROM dorixonalar")
                        return [dict(row) for row in cur.fetchall()]
            except Exception as e:
                logger.error(f"Database error in get_all_pharmacies: {e}")
                return []

   

    def execute_query(self, query: str) -> List[Dict]:
        """Query bajarish va natijani ro'yxat ko'rinishida qaytarish"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(query)
                    return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Database error in execute_query: {e}")
            return []
        




    def get_connection(self):
        return psycopg2.connect(**self.config)
    

    def check_employee(self, phone_number: str) -> Optional[Dict]:
        """Telefon raqami bo'yicha xodimni tekshirish"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    print(f"🔍 DEBUG: Qidirilayotgan raqam: {phone_number}")
                    
                    # Telefon raqamini turli formatlarda qidirish
                    search_variants = [
                        phone_number,  # +998901234567
                        phone_number[1:] if phone_number.startswith('+') else phone_number,  # 998901234567
                        phone_number[4:] if phone_number.startswith('+998') else phone_number,  # 901234567
                        phone_number.replace('+', ''),  # 998901234567
                        phone_number.replace('+998', ''),  # 901234567
                        phone_number.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')  # Tozalash
                    ]
                    
                    print(f"🔍 DEBUG: Qidiruv variantlari: {search_variants}")
                    
                    # Bazadagi barcha telefon raqamlarni ko'rish (debug uchun)
                    cur.execute("SELECT ism_familiya, telefon_raqam FROM xodimlar LIMIT 5")
                    all_phones = cur.fetchall()
                    print(f"🔍 DEBUG: Bazadagi telefon raqamlar namunasi: {[dict(row) for row in all_phones]}")
                    
                    # Har bir variant bo'yicha qidirish
                    for variant in search_variants:
                        cur.execute(
                            "SELECT * FROM xodimlar WHERE telefon_raqam = %s OR telefon_raqam LIKE %s",
                            (variant, f"%{variant}%")
                        )
                        result = cur.fetchone()
                        if result:
                            print(f"🔍 DEBUG: Topildi! Variant: {variant}")
                            return dict(result)
                    
                    print("🔍 DEBUG: Hech narsa topilmadi!")
                    return None
                    
        except Exception as e:
            logger.error(f"Database error in check_employee: {e}")
            print(f"🔍 DEBUG: Baza xatosi: {e}")
            return None

    def get_pharmacy_by_inn(self, inn: str) -> Optional[Dict]:
        """INN bo'yicha dorixonani topish"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(
                        "SELECT * FROM dorixonalar WHERE inn = %s",
                        (inn,)
                    )
                    result = cur.fetchone()
                    return dict(result) if result else None
        except Exception as e:
            logger.error(f"Database error: {e}")
            return None
    
    def search_pharmacy_by_name(self, name: str) -> List[Dict]:
        """Nom bo'yicha dorixonalarni qidirish"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(
                        "SELECT * FROM dorixonalar WHERE LOWER(dorixona_nomi) LIKE LOWER(%s)",
                        (f"%{name}%",)
                    )
                    return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Database error: {e}")
            return []
    
    def get_all_medicines(self) -> List[Dict]:
        """Barcha dorilarni olish (rasm va ma'lumot bilan)"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT id, dori_nomi, narxi, ikpu, image_file_id, info FROM dorilar ORDER BY dori_nomi")
                    return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Database error: {e}")
            return []

    def get_medicines_for_price_list(self):
        """Narxlar ro'yxati uchun faqat kerakli ma'lumotlarni olish"""
        query = "SELECT dori_nomi, narxi, ikpu FROM dorilar ORDER BY dori_nomi"
        return self.execute_query(query)
    
    def get_price_list_image(self) -> Optional[str]:
        """Narxlar ro'yxati uchun yagona rasm (file_id) ni olish"""
        query = "SELECT rasm FROM dorilar WHERE rasm IS NOT NULL LIMIT 1"
        result = self.execute_query(query)
        return result[0]['rasm'] if result else None

        





# Global o'zgaruvchilar
db_manager = DatabaseManager(DB_CONFIG)
user_sessions = {}  # Foydalanuvchi sessiyalarini saqlash

class UserSession:
     def __init__(self):
        self.is_authorized = False
        self.phone_number = None
        self.employee_info = None
        self.day_started = False
        self.current_pharmacy = None
        self.current_order = []
        self.order_total = 0
        self.discount_percentage = 0
        self.state = "idle"
        self.permanently_authorized = False



import re
import unicodedata
from difflib import SequenceMatcher


class PharmacySearchSystem:
    def __init__(self):
        # Lotin-Kirill transliteratsiya jadvali
        self.cyrillic_to_latin = {
            'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
            'ё': 'yo', 'ж': 'j', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
            'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
            'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'x', 'ц': 'ts',
            'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ъ': '', 'ы': 'i', 'ь': '',
            'э': 'e', 'ю': 'yu', 'я': 'ya',
            'қ': 'q', 'ғ': 'g', 'ҳ': 'h', 'ў': 'o'
        }
        
        self.latin_to_cyrillic = {v: k for k, v in self.cyrillic_to_latin.items() if v}
        
        # PHARM/FARM variantlari
        self.pharm_variants = {
            'pharm', 'pхarm', 'фарм', 'пхарм', 'farm', 'fарм', 'пharм'
        }
        
        # Umumiy qisqartmalar
        self.common_abbreviations = {
            'mchj': ['mchj', 'мчж', 'масъулияти чекланган жамият'],
            'qmj': ['qmj', 'қмж', 'қўшма мулк жамияти'],
            'tj': ['tj', 'тж', 'тадбиркорлик жамияти'],
            'ooo': ['ooo', 'ооо', 'открытое акционерное общество'],
            'ltd': ['ltd', 'лтд', 'limited'],
            'co': ['co', 'ко', 'company']
        }

    def normalize_text(self, text):
        """Matnni normallashtirishni"""
        if not text:
            return ""
        
        # Kichik harflarga o'tkazish
        text = text.lower().strip()
        
        # Maxsus belgilarni olib tashlash
        text = re.sub(r'[^\w\s]', ' ', text)
        
        # Ko'p bo'shliqlarni bitta bo'shliqqa almashtirish
        text = re.sub(r'\s+', ' ', text)
        
        # Unicode normalizatsiya
        text = unicodedata.normalize('NFKD', text)
        
        return text

    def transliterate_to_latin(self, text):
        """Kirill matnini lotinga o'girish"""
        result = ""
        for char in text.lower():
            result += self.cyrillic_to_latin.get(char, char)
        return result

    def transliterate_to_cyrillic(self, text):
        """Lotin matnini kirillga o'girish (oddiy hollarda)"""
        result = text
        # Uzun kombinatsiyalarni birinchi navbatda almashtirish
        for lat, cyr in sorted(self.latin_to_cyrillic.items(), key=len, reverse=True):
            if len(lat) > 1:
                result = result.replace(lat, cyr)
        
        # Yakka harflarni almashtirish
        for lat, cyr in self.latin_to_cyrillic.items():
            if len(lat) == 1:
                result = result.replace(lat, cyr)
        
        return result

    def normalize_pharm_word(self, text):
        """PHARM/FARM so'zlarini normallashtiish"""
        words = text.split()
        normalized_words = []
        
        for word in words:
            word_lower = word.lower()
            if word_lower in self.pharm_variants:
                normalized_words.append('pharm')
            else:
                normalized_words.append(word)
        
        return ' '.join(normalized_words)

    def expand_abbreviations(self, text):
        """Qisqartmalarni kengaytirish"""
        words = text.split()
        expanded_variants = [text]  # Asl matn ham qo'shiladi
        
        for abbr, variants in self.common_abbreviations.items():
            for i, word in enumerate(words):
                if word.lower() in variants:
                    # Har bir variant uchun yangi matn yaratish
                    for variant in variants:
                        new_words = words.copy()
                        new_words[i] = variant
                        expanded_variants.append(' '.join(new_words))
        
        return list(set(expanded_variants))  # Takrorlarni olib tashlash

    def generate_search_variants(self, query):
        """Qidiruv so'rovi uchun barcha variantlarni yaratish"""
        if not query:
            return []
        
        variants = set()
        
        # Asl so'rovni normallash
        normalized_query = self.normalize_text(query)
        variants.add(normalized_query)
        
        # PHARM so'zlarini normallash
        pharm_normalized = self.normalize_pharm_word(normalized_query)
        variants.add(pharm_normalized)
        
        # Transliteratsiya variantlari
        latin_variant = self.transliterate_to_latin(normalized_query)
        cyrillic_variant = self.transliterate_to_cyrillic(normalized_query)
        variants.add(latin_variant)
        variants.add(cyrillic_variant)
        
        # Qisqartmalar bilan kengaytirish
        for variant in list(variants):
            expanded = self.expand_abbreviations(variant)
            variants.update(expanded)
        
        # Bo'sh qatorlarni olib tashlash
        variants = {v for v in variants if v.strip()}
        
        return list(variants)

    def calculate_similarity(self, text1, text2):
        """Ikki matn orasidagi o'xshashlikni hisoblash"""
        if not text1 or not text2:
            return 0.0
        
        # Asosiy o'xshashlik
        similarity = SequenceMatcher(None, text1.lower(), text2.lower()).ratio()
        
        # So'zlar bo'yicha o'xshashlik
        words1 = set(text1.lower().split())
        words2 = set(text2.lower().split())
        
        if words1 and words2:
            word_similarity = len(words1.intersection(words2)) / len(words1.union(words2))
            similarity = max(similarity, word_similarity)
        
        return similarity

    def contains_search_term(self, text, search_term):
        """Matn ichida qidiruv terminini tekshirish"""
        if not text or not search_term:
            return False
        
        text_lower = text.lower()
        search_lower = search_term.lower()
        
        # To'g'ridan-to'g'ri mavjudlik
        if search_lower in text_lower:
            return True
        
        # Transliteratsiya orqali tekshirish
        text_latin = self.transliterate_to_latin(text_lower)
        text_cyrillic = self.transliterate_to_cyrillic(text_lower)
        search_latin = self.transliterate_to_latin(search_lower)
        search_cyrillic = self.transliterate_to_cyrillic(search_lower)
        
        # Barcha kombinatsiyalarni tekshirish
        combinations = [
            (search_lower, text_lower),
            (search_latin, text_latin),
            (search_cyrillic, text_cyrillic),
            (search_lower, text_latin),
            (search_lower, text_cyrillic),
            (search_latin, text_lower),
            (search_cyrillic, text_lower)
        ]
        
        for search_var, text_var in combinations:
            if search_var in text_var:
                return True
        
        return False

    def search_pharmacies(self, query, pharmacies, min_similarity=0.1):
        """Dorixonalarni qidirish"""
        if not query.strip():
            return []
        
        # INN raqam bo'yicha qidirish
        if query.strip().isdigit():
            exact_matches = [p for p in pharmacies if p.get('inn') == query.strip()]
            if exact_matches:
                return exact_matches
        
        results = []
        query_normalized = self.normalize_text(query)
        
        for pharmacy in pharmacies:
            pharmacy_name = pharmacy.get('dorixona_nomi', '')
            pharmacy_address = pharmacy.get('manzil', '')
            pharmacy_owner = pharmacy.get('dorixona_egasi', '')
            
            max_score = 0.0
            match_field = ""
            
            # 1. To'g'ridan-to'g'ri qidirish (eng yuqori ball)
            if self.contains_search_term(pharmacy_name, query):
                max_score = max(max_score, 1.0)
                match_field = "nom (to'liq)"
            
            if self.contains_search_term(pharmacy_address, query):
                max_score = max(max_score, 0.9)
                match_field = "manzil (to'liq)"
            
            if self.contains_search_term(pharmacy_owner, query):
                max_score = max(max_score, 0.8)
                match_field = "egasi (to'liq)"
            
            # 2. So'zlar bo'yicha qidirish
            query_words = query_normalized.split()
            for word in query_words:
                if len(word) >= 2:  # Kamida 2 harfli so'zlarni qidirish
                    if self.contains_search_term(pharmacy_name, word):
                        max_score = max(max_score, 0.7)
                        match_field = "nom (qisman)"
                    
                    if self.contains_search_term(pharmacy_address, word):
                        max_score = max(max_score, 0.6)
                        match_field = "manzil (qisman)"
                    
                    if self.contains_search_term(pharmacy_owner, word):
                        max_score = max(max_score, 0.5)
                        match_field = "egasi (qisman)"
            
            # 3. O'xshashlik tekshiruvi (eng past ball)
            if max_score == 0:
                # Qidiruv variantlarini yaratish
                search_variants = self.generate_search_variants(query)
                name_variants = self.generate_search_variants(pharmacy_name)
                
                for search_var in search_variants:
                    for name_var in name_variants:
                        sim = self.calculate_similarity(search_var, name_var)
                        if sim > max_score and sim >= min_similarity:
                            max_score = sim * 0.4  # O'xshashlik ballini kamaytirish
                            match_field = "nom (o'xshash)"
            
            # Natijaga qo'shish
            if max_score > 0:
                results.append({
                    'pharmacy': pharmacy,
                    'score': max_score,
                    'match_field': match_field
                })
        
        # Natijalarni ball bo'yicha saralash
        results.sort(key=lambda x: x['score'], reverse=True)
        
        return [r['pharmacy'] for r in results]

# Qidiruv tizimini ishlatish uchun funksiya
def search_pharmacy_advanced(query, pharmacies, min_similarity=0.1):
    """Rivojlangan dorixona qidiruv funksiyasi"""
    search_system = PharmacySearchSystem()
    return search_system.search_pharmacies(query, pharmacies, min_similarity)





AUTHORIZED_USERS = {}  # {user_id: {'phone': phone, 'employee_info': employee_data}}






def get_medicines_keyboard(medicines, page=0, per_page=10):
    """Dorilar ro'yxati uchun keyboard"""
    keyboard = []
    start_idx = page * per_page
    end_idx = start_idx + per_page
    
    # Dorilarni sahifalash
    current_medicines = medicines[start_idx:end_idx]
    
    # Har bir dori uchun tugma (faqat nom va narx)
    for medicine in current_medicines:
        keyboard.append([InlineKeyboardButton(
            f"💊 {medicine['dori_nomi']} - {medicine['narxi']:,} so'm",
            callback_data=f"select_medicine_{medicine['id']}"
        )])
    
    # Sahifalash tugmalari
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("⬅️ Олдинги", callback_data=f"page_{page-1}"))
    if end_idx < len(medicines):
        nav_buttons.append(InlineKeyboardButton("Кейинги ➡️", callback_data=f"page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Buyurtmani yakunlash tugmasi (agar buyurtma mavjud bo'lsa)
    keyboard.append([InlineKeyboardButton("✅ Буюрмани якунлаш", callback_data="finish_order")])
    keyboard.append([InlineKeyboardButton("❌ Бекор қилиш", callback_data="cancel_order")])
    
    return InlineKeyboardMarkup(keyboard)

def get_edit_quantity_keyboard(medicine_id):
    """Miqdorni tahrirlash uchun keyboard"""
    keyboard = [
        [InlineKeyboardButton("✏️ Таҳрирлаш", callback_data=f"edit_quantity_{medicine_id}")],
        [InlineKeyboardButton("🔙 Орқага", callback_data="back_to_medicines")]
    ]
    return InlineKeyboardMarkup(keyboard)



# 1. get_start_day_keyboard() funksiyasini o'zgartirish
def get_start_day_keyboard():
    """Kunni boshlash klaviaturasi"""
    keyboard = [
        [KeyboardButton("🌅 Иш кунини бошлаш", request_location=True)]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_pharmacy_confirm_keyboard():
    """Dorixona tasdiqlash klaviaturasi"""
    keyboard = [
        [InlineKeyboardButton("✅ Ҳa, тўғри", callback_data="confirm_pharmacy_yes")],
        [InlineKeyboardButton("❌ Йўқ, қайтадан", callback_data="confirm_pharmacy_no")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_location_request_keyboard():
    """Geolokatsiya so'rash klaviaturasi"""
    keyboard = [
        [KeyboardButton("📍 Геолокация юбориш", request_location=True)]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)



# 3. get_main_menu_keyboard() funksiyasini o'zgartirish
def get_main_menu_keyboard():
    """Asosiy menyu klaviaturasi"""
    keyboard = [
        [KeyboardButton("🏥 Дорихона"), KeyboardButton("📋 Буюртма олиш")],
        [KeyboardButton("💰 Нархлар рўйхати"), KeyboardButton("🔚 Иш кунининг охири", request_location=True)],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)








def get_payment_keyboard():
    """To'lov tanlash klaviaturasi"""
    keyboard = [
        [InlineKeyboardButton(" 100% тўлов", callback_data="payment_100")],
        [InlineKeyboardButton(" 100% дан кам миқдорда тўлов", callback_data="payment_50")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_confirm_keyboard():
    """Tasdiqlash klaviaturasi"""
    keyboard = [
        [InlineKeyboardButton("✅ Тасдиқлаш", callback_data="confirm_yes")],
        [InlineKeyboardButton("❌ Бекор қилиш", callback_data="confirm_no")]
    ]
    return InlineKeyboardMarkup(keyboard)



async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Telefon kontaktini qayta ishlash"""
    user_id = update.effective_user.id
    contact = update.message.contact
    
    # Kontakt foydalanuvchining o'zinikiligini tekshirish
    if contact.user_id != user_id:
        await update.message.reply_text(
            "❌ Фақат ўз телефон рақамингизни юборинг!",
            reply_markup=get_phone_request_keyboard()
        )
        return
    
    phone_number = contact.phone_number
    print(f"🔍 DEBUG: Qabul qilingan telefon raqam: {phone_number}")
    
    # Telefon raqamini tozalash (+998 dan boshlanishi kerak)
    if not phone_number.startswith('+'):
        phone_number = '+' + phone_number
    
    print(f"🔍 DEBUG: Tozalangan telefon raqam: {phone_number}")
    
    # Ma'lumotlar bazasidan xodimni tekshirish
    employee = db_manager.check_employee(phone_number)
    print(f"🔍 DEBUG: Bazadan natija: {employee}")
    
    if not employee:
        await update.message.reply_text(
            f"❌ Телефон рақамингиз ({phone_number}) рўйхатга олинмаган.\n\n"
            "📞 Администратор билан боғланинг: +998 99 830 23 30 / Telegram: @gynomedixmavluda\n"
            "Қайтадан уриниш учун /start босинг."
        )
        return
    
    # Muvaffaqiyatli avtorizatsiya
    if user_id not in user_sessions:
        user_sessions[user_id] = UserSession()
    
    session = user_sessions[user_id]
    session.is_authorized = True
    session.permanently_authorized = True  # YANGI
    session.phone_number = phone_number
    session.employee_info = employee
    
    # YANGI: Doimiy avtorizatsiyani saqlash
    AUTHORIZED_USERS[user_id] = {
        'phone': phone_number,
        'employee_info': employee
    }
    
    employee_name = employee.get('ism_familiya', 'Noma\'lum')
    # employee_position = employee.get('lavozim', 'Noma\'lum')
    
    await update.message.reply_text(
        f"✅ Муваффақиятли авторизация қилиндингиз!\n\n"
        f"👤 Исм: {employee_name}\n"
        # f"💼 Лавозим: {employee_position}\n"
        f"📞 Телефон: {phone_number}\n\n"
        "Иш кунини бошлаш учун тугмани босинг:",
        reply_markup=get_start_day_keyboard()
    )


# Yangi keyboard funksiyasi - telefon raqam so'rash uchun
def get_phone_request_keyboard():
    """Telefon raqam so'rash klaviaturasi"""
    keyboard = [
        [KeyboardButton("📞 Telefon raqamni yuborish", request_contact=True)]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)











async def handle_pharmacy_search_advanced(update, context, query):
    """Rivojlangan dorixona qidirish"""
    user_id = update.effective_user.id
    session = user_sessions[user_id]
    
    # INN raqam bo'yicha to'g'ridan-to'g'ri qidirish
    if query.strip().isdigit():
        pharmacy = db_manager.get_pharmacy_by_inn(query.strip())
        if pharmacy:
            session.current_pharmacy = pharmacy
            
            pharmacy_info = f"🏥 *{pharmacy['dorixona_nomi']}*\n"
            pharmacy_info += f"📍 Манзил: {pharmacy['manzil']}\n"
            pharmacy_info += f"🏢 ИНН: {pharmacy['inn']}\n"
            pharmacy_info += f"👤 Эгаси: {pharmacy['dorixona_egasi']}\n"
            pharmacy_info += f"📞: {pharmacy['kontrakt']}\n"
            
            await update.message.reply_text(
                pharmacy_info + "\n\n❓ Бу дорихона тўғрими?",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_pharmacy_confirm_keyboard()
            )
            return
        else:
            await update.message.reply_text(
                "❌ Бундай ИНН рақамли дорихона топилмади!\n"
                "Қайтадан киритинг ёки ном бўйича қидиринг:"
            )
            return
    
    # Nom bo'yicha rivojlangan qidirish
    all_pharmacies = db_manager.get_all_pharmacies()  # Barcha dorixonalarni olish
    found_pharmacies = search_pharmacy_advanced(query, all_pharmacies, min_similarity=0.1)
    
        
    if found_pharmacies:
        if len(found_pharmacies) == 1:
            pharmacy = found_pharmacies[0]
            session.current_pharmacy = pharmacy
            
            pharmacy_info = f"🏥 <b>{escape(pharmacy['dorixona_nomi'])}</b>\n"
            pharmacy_info += f"📍 Манзил: {escape(pharmacy['manzil'])}\n"
            pharmacy_info += f"🏢 ИНН: <code>{escape(pharmacy['inn'])}</code>\n"
            pharmacy_info += f"👤 Эгаси: {escape(pharmacy['dorixona_egasi'])}\n"
            pharmacy_info += f"📞: {escape(pharmacy['kontrakt'])}\n"
            
            await update.message.reply_text(
                pharmacy_info + "\n\n❓ Бу дорихона тўғрими?",
                parse_mode=ParseMode.HTML,
                reply_markup=get_pharmacy_confirm_keyboard()
            )
        else:
            result = "🔍 <b>Топилган дорихоналар:</b>\n\n"
            for i, pharmacy in enumerate(found_pharmacies[:10], 1):
                result += f"{i}. <b>{escape(pharmacy['dorixona_nomi'])}</b>\n"
                result += f"   ИНН: <code>{escape(pharmacy['inn'])}</code>\n"
                result += f"   📍 {escape(pharmacy['manzil'])}\n"
                if pharmacy.get('kontrakt'):
                    result += f"   📞 {escape(pharmacy['kontrakt'])}\n"
                result += "\n"
            
            result += "📝 <b>ИНН рақамини киритинг:</b>"
            await update.message.reply_text(result, parse_mode=ParseMode.HTML)



    else:
        await update.message.reply_text(
            f"❌ *'{query}'* бўйича дорихона топилмади!\n\n"
            "💡 *Қидириш бўйича маслахатлар:*\n"
            "• Дорихона номини тўлиқ киритинг\n"
            "• ИНН рақамини киритиб кўринг\n"
            "• Lotin ёки Кирилл харфларида ёзиб кўринг\n"
            "• PHARM, ФАРМ каби қисқартмаларни ишлатинг\n\n"
            "Қайтадан уриниб кўринг:",
            parse_mode=ParseMode.MARKDOWN
        )



















# 3. O'zgartirilgan start funksiyasi
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bot boshlash buyrug'i"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions:
        user_sessions[user_id] = UserSession()
    
    session = user_sessions[user_id]
    
    # YANGI: Avval doimiy avtorizatsiya borligini tekshirish
    if user_id in AUTHORIZED_USERS:
        # Agar doimiy avtorizatsiya mavjud bo'lsa, uni tiklaymiz
        session.is_authorized = True
        session.permanently_authorized = True
        session.phone_number = AUTHORIZED_USERS[user_id]['phone']
        session.employee_info = AUTHORIZED_USERS[user_id]['employee_info']
        
        employee_name = session.employee_info.get('ism_familiya', 'Noma\'lum')
        await update.message.reply_text(
            f"👋 Янги иш кунини бошлаганингиздан ҳурсандмиз, хуш келибсиз, {employee_name}!\n\n"
            "Иш кунини бошлаш учун тугмани босинг:",
            reply_markup=get_start_day_keyboard()
        )
        return
    
    # Agar eski usulda avtorizatsiya qilingan bo'lsa (lekin doimiy emas)
    if session.is_authorized and not session.permanently_authorized:
        await update.message.reply_text(
            "Сиз аллақачон тизимга киргансиз!\n"
            "Иш кунини бошлаш учун тугмани босинг:",
            reply_markup=get_start_day_keyboard()
        )
        return
    
    # Telefon raqam so'rash (faqat yangi foydalanuvchilar uchun)
    await update.message.reply_text(
        "👋 Ассалому алайкум!\n\n"
        "🔐 Тизимга кириш учун телефон рақамингизни юборинг.\n"
        "Қуйидаги тугмани босинг:",
        reply_markup=get_phone_request_keyboard()
    )


async def handle_video_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Video xabar qayta ishlash"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions or not user_sessions[user_id].is_authorized:
        return
    
    session = user_sessions[user_id]
    video_note = update.message.video_note
    
    if video_note.duration < 10:
        await update.message.reply_text(
            "⚠️ Видео хабар камида 10 сония бўлиши керак! Қайта юборинг."
        )
        return
    
    if not session.day_started and hasattr(session, 'location_received'):
        # Kun boshlash
        session.day_started = True
        await update.message.reply_text(
            "✅ Иш куни муваффақиятли бошланди!\n"
            "Асосий менюдан керакли бўлимни танланг:",
            reply_markup=get_main_menu_keyboard()
        )
    elif session.day_started:
        # Kun tugashi - FAQAT ish kunini tugatish, avtorizatsiyani saqlab qolish
        session.day_started = False
        session.current_pharmacy = None
        session.current_order = []
        session.order_total = 0
        session.discount_percentage = 0
        session.state = "idle"
        # DIQQAT: is_authorized va permanently_authorized o'zgartirilmaydi!
        
        employee_name = session.employee_info.get('ism_familiya', 'Noma\'lum')
        await update.message.reply_text(
            f"✅ Иш куни якунланди, {employee_name}!\n"
            "Хайр-саломат бўлинг!\n"
            "Савол ва мурожаатлар учун администратор билан боғланинг: +998 99 830 23 30 / Telegram: @gynomedixmavluda\n\n"
            "🔄 Эртага иш кунини бошлаш учун /start босинг.",
            reply_markup=None
        )












async def reset_auth(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Avtorizatsiyani qayta o'rnatish (test uchun)"""
    user_id = update.effective_user.id
    
    # AUTHORIZED_USERS dan o'chirish
    if user_id in AUTHORIZED_USERS:
        del AUTHORIZED_USERS[user_id]
    
    # Sessionni tozalash
    if user_id in user_sessions:
        user_sessions[user_id] = UserSession()
    
    await update.message.reply_text(
        "🔄 Авторизация қайта ўрнатилди!\n"
        "Қайтадан /start босинг."
    )




async def handle_pharmacy_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Dorixona qidirish"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions or not user_sessions[user_id].day_started:
        await update.message.reply_text("Аввал иш кунини бошланг!")
        return
    
    session = user_sessions[user_id]
    session.state = "pharmacy_search"
    
    await update.message.reply_text(
        "🔍 Дорихонани қидириш учун ИНН рақамини ёки дорихона номини киритинг:"
    )


async def handle_order_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Buyurtma olishni boshlash"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions or not user_sessions[user_id].day_started:
        await update.message.reply_text("Аввал иш кунини бошланг!")
        return
    
    session = user_sessions[user_id]
    
    if not session.current_pharmacy:
        await update.message.reply_text(
            "❌ Аввал дорихонани танланг!\n"
            "🏥 Дорихона тугмасини босинг."
        )
        return
    
    medicines = db_manager.get_all_medicines()
    if not medicines:
        await update.message.reply_text("❌ Дорилар рўйхати бўш!")
        return
    
    session.state = "order_taking"
    session.current_order = []
    session.order_total = 0
    
    # Dorilar ro'yxatini tugmacha ko'rinishida yuborish
    await update.message.reply_text(
        f"🏥 **{session.current_pharmacy['dorixona_nomi']}** учун буюртма\n\n"
        "💊 Қуйидаги дорилардан танланг:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=get_medicines_keyboard(medicines)
    )

async def create_specification(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Spetsifikatsiya yaratish"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions or not user_sessions[user_id].day_started:
        await update.message.reply_text("Аввал иш кунини бошланг!")
        return
    
    session = user_sessions[user_id]
    
    if not session.current_pharmacy or not session.current_order:
        await update.message.reply_text(
            "❌ Аввал дорихона танланг ва буюртма қилинг!"
        )
        return
    
    # Excel fayl yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Спецификация"
    
    # Sarlavhalar
    headers = ["Поставщик", "Покупатель", "Товар", "ИКПУ", "Количество", "Цена", "Сумма"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Ma'lumotlarni to'ldirish
    supplier_info = 'MCHJ "GYNOMEDIX"\nАДРЕС: Toshkent shaxri Chilonzor tumani. Dumbirobod 4 tor kuchasi 23/2\nТЕЛ.: 99 830-23-30\nИНН: 311818897'
    
    pharmacy = session.current_pharmacy
    buyer_info = f"{pharmacy['dorixona_nomi']}\n{pharmacy['manzil']}\nИНН: {pharmacy['inn']}"
    
    row = 2
    for item in session.current_order:
        ws.cell(row=row, column=1, value=supplier_info)
        ws.cell(row=row, column=2, value=buyer_info)
        ws.cell(row=row, column=3, value=item['name'])
        ws.cell(row=row, column=4, value=item['ikpu'])
        ws.cell(row=row, column=5, value=item['quantity'])
        ws.cell(row=row, column=6, value=item['price'])
        ws.cell(row=row, column=7, value=item['total'])
        row += 1
    
    # Faylni saqlash
    filename = f"specification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    # Faylni yuborish
    with open(filename, 'rb') as file:
        await update.message.reply_document(
            document=file,
            filename=filename,
            caption="📋 Спецификация тайёр!"
        )
    
    # Faylni o'chirish
    os.remove(filename)




async def send_price_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Narxlar ro'yxatini rasm ko'rinishida yuborish"""
    user_id = update.effective_user.id

    if user_id not in user_sessions or not user_sessions[user_id].day_started:
        await update.message.reply_text("Аввал иш кунини бошланг!")
        return

    image_file_id = db_manager.get_price_list_image()
    if not image_file_id:
        await update.message.reply_text("❌ Нархлар рўйхати расми топилмади!")
        return

    await update.message.reply_photo(
        caption="💰 *Нархлар рўйхати*",
        photo=image_file_id,
        parse_mode=ParseMode.MARKDOWN
    )




# 4. handle_text_message() funksiyasidagi tegishli qismlarni o'zgartirish
async def handle_text_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Matnli xabarlarni qayta ishlash"""
    user_id = update.effective_user.id
    text = update.message.text
    
    if user_id not in user_sessions:
        await start(update, context)
        return
    
    
    session = user_sessions[user_id]


    
    if user_id in ADMIN_IDS and context.user_data.get('admin_state'):
            await admin_handle_text_message(update, context)
            return
    

    # Avtorizatsiya tekshiruvi
    if not session.is_authorized:
        await update.message.reply_text(
            "❌ Аввал телефон рақамингизни юборинг!\n"
            "/start тугмасини босинг.",
            reply_markup=get_phone_request_keyboard()
        )
        return
    
    if not session.day_started:
        if text == "🌅 Иш кунини бошлаш":
            # Geolokatsiya avtomatik yuboriladi, foydalanuvchiga xabar berish
            await update.message.reply_text(
                "📍 Геолокациянгиз автоматик юборилди!\n"
                "Энди камида 10 сония давомида видео хабар юборинг."
            )
        return

    # Asosiy menyu tugmalari
    if text == "🏥 Дорихона":
        await handle_pharmacy_search(update, context)
    elif text == "📋 Буюртма олиш":
        await handle_order_start(update, context)
    elif text == "💰 Нархлар рўйхати":
        await send_price_list(update, context)
    elif text == "🔚 Иш кунининг охири":
        # Geolokatsiya avtomatik yuboriladi, foydalanuvchiga xabar berish
        await update.message.reply_text(
            "📍 Геолокациянгиз автоматик юборилди!\n"
            "Энди камида 10 сония давомида видео хабар юборинг."
        )
    
    # Dorixona qidirish qismida (handle_text_message ichida):
    elif session.state == "pharmacy_search":
        await handle_pharmacy_search_advanced(update, context, text)
       
    elif session.state == "order_taking":
        try:
            # Buyurtmani tugatish
            if text.lower() in ['tugadi', 'tugatish', 'yakunlash']:
                if session.current_order:
                    session.state = "payment_selection"
                    await update.message.reply_text(
                        f"📋 *Буюртма хулосаси:*\n"
                        f"🏥 Дорихона: {session.current_pharmacy['dorixona_nomi']}\n"
                        f"💰 Умумий сумма: {session.order_total:,} so'm\n\n"
                        "Тўлов турини танланг:",
                        parse_mode=ParseMode.MARKDOWN,
                        reply_markup=get_payment_keyboard()
                    )
                else:
                    await update.message.reply_text("❌ Буюртма бўш!")
                return
                
            # Buyurtma formatini tahlil qilish: "dori_nomi miqdor"
            parts = text.rsplit(' ', 1)
            if len(parts) != 2:
                await update.message.reply_text(
                    "❌ Нотўғри формат!\n"
                    "Тўғри формат: АРГИСАН сироп 5"
                )
                return
            
            medicine_name, quantity_str = parts
            quantity = int(quantity_str)
            
            # Dorini topish
            medicines = db_manager.get_all_medicines()
            medicine = None
            for med in medicines:
                if medicine_name.lower() in med['dori_nomi'].lower():
                    medicine = med
                    break
            
            if not medicine:
                await update.message.reply_text(
                    f"❌ '{medicine_name}' номи билан дори топилмади!\n"
                    "Рўйхатдаги тўғри номни киритинг."
                )
                return
            
            # Buyurtmaga qo'shish
            item_total = medicine['narxi'] * quantity
            order_item = {
                'name': medicine['dori_nomi'],
                'price': medicine['narxi'],
                'quantity': quantity,
                'total': item_total,
                'ikpu': medicine['ikpu']
            }
            
            session.current_order.append(order_item)
            session.order_total += item_total
            
            await update.message.reply_text(
                f"✅ Қўшилди: {medicine['dori_nomi']} x {quantity}\n"
                f"💰 Сумма: {item_total:,} so'm\n"
                f"📊 Умумий сумма: {session.order_total:,} so'm\n\n"
                "Яна қўшиш учун киритинг ёки 'тугади' деб ёзинг:"
            )
            
        except ValueError:
            await update.message.reply_text(
                "❌ Миқдорни тўғри киритинг (рақам бўлиши керак)!"
            )
        except Exception as e:
            logger.error(f"Order processing error: {e}")
            await update.message.reply_text("❌ Хатолик юз берди!")

    elif session.state == "quantity_input":
        try:
            quantity = int(text)
            if quantity <= 0:
                await update.message.reply_text("❌ Миқдор 0 дан катта бўлиши керак!")
                return
            
            medicine = context.user_data.get('selected_medicine')
            if not medicine:
                await update.message.reply_text("❌ Хатолик юз берди, қайтадан уриниб кўринг!")
                return
            
            # Buyurtmaga qo'shish
            item_total = medicine['narxi'] * quantity
            order_item = {
                'id': medicine['id'],
                'name': medicine['dori_nomi'],
                'price': medicine['narxi'],
                'quantity': quantity,
                'total': item_total,
                'ikpu': medicine['ikpu']
            }
            
            session.current_order.append(order_item)
            session.order_total += item_total
            session.state = "order_taking"
            
            await update.message.reply_text(
                f"✅ **{medicine['dori_nomi']}** учун **{quantity}** та буюртма қабул қилинди\n"
                f"💰 Сумма: {item_total:,} сўм\n"
                f"📊 Умумий сумма: {session.order_total:,} сўм",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_edit_quantity_keyboard(medicine['id'])
            )
            
        except ValueError:
            await update.message.reply_text("❌ Миқдорни тўғри киритинг (рақам бўлиши керак)!")
    
    # YANGI: Miqdorni tahrirlash
    elif session.state == "edit_quantity":
        try:
            new_quantity = int(text)
            if new_quantity <= 0:
                await update.message.reply_text("❌ Миқдор 0 дан катта бўлиши керак!")
                return
            
            medicine = context.user_data.get('edit_medicine')
            if not medicine:
                await update.message.reply_text("❌ Хатолик юз берди!")
                return
            
            # Buyurtmadan topib yangilash
            for item in session.current_order:
                if item['id'] == medicine['id']:
                    old_total = item['total']
                    item['quantity'] = new_quantity
                    item['total'] = medicine['narxi'] * new_quantity
                    session.order_total = session.order_total - old_total + item['total']
                    break
            
            session.state = "order_taking"
            
            await update.message.reply_text(
                f"✅ **{medicine['dori_nomi']}** миқдори **{new_quantity}** га  ўзгартирилди\n"
                f"💰 Yangi summa: {medicine['narxi'] * new_quantity:,} сўм\n"
                f"📊 Umumiy summa: {session.order_total:,} сўм",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_edit_quantity_keyboard(medicine['id'])
            )
            
        except ValueError:
            await update.message.reply_text("❌ Миқдорни тўғри киритинг (рақам бўлиши керак)!")







async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback query larni qayta ishlash"""
    query = update.callback_query
    user_id = query.from_user.id
    data = query.data
    
    await query.answer()
    
    if user_id not in user_sessions:
        return
    
    session = user_sessions[user_id]

    # Dorixona tasdiqlash
    if data.startswith("confirm_pharmacy"):
        if data == "confirm_pharmacy_yes":
            session.state = "waiting_pharmacy_location"  # YANGI holat
            await query.edit_message_text(
                "✅ Дорихона тасдиқланди!\n"
                "📍 Геолокациянгизни юборинг:",
                reply_markup=None
            )
            await query.message.reply_text(
                "Қуйидаги тугмани босиб геолокациянгизни юборинг:",
                reply_markup=get_location_request_keyboard()
            )
        else:
            session.current_pharmacy = None
            session.state = "pharmacy_search"
            await query.edit_message_text(
                "❌ Дорихона бекор қилинди!\n"
                "Қайтадан ИНН рақами ёки ном киритинг:",
                reply_markup=None
            )
        return

    # YANGI: Dori tanlash (handle_callback_query ichida bu qismni almashtiring)
    if data.startswith("select_medicine_"):
        medicine_id = int(data.split("_")[2])
        medicines = db_manager.get_all_medicines()
        medicine = next((m for m in medicines if m['id'] == medicine_id), None)
        
        if medicine:
            context.user_data['selected_medicine'] = medicine
            session.state = "quantity_input"
            
            # Avval rasmni yuborish (agar mavjud bo'lsa)
            if medicine.get('image_file_id'):
                try:
                    # Rasm va ma'lumot bilan birga yuborish
                    caption = f"💊 **{medicine['dori_nomi']}**\n"
                    caption += f"💰 Нарх: {medicine['narxi']:,} сўм\n"
                    caption += f"🏷️ ИКПУ: {medicine['ikpu']}\n\n"
                    
                    # Info mavjud bo'lsa qo'shish
                    if medicine.get('info'):
                        caption += f"ℹ️ **Маълумот:**\n{medicine['info']}\n\n"
                    
                    caption += "📝 Дори учун миқдор киритинг:"
                    
                    await query.message.reply_photo(
                        photo=medicine['image_file_id'],
                        caption=caption,
                        parse_mode=ParseMode.MARKDOWN
                    )
                    
                    # Eski xabarni o'chirish
                    await query.delete_message()
                    
                except Exception as e:
                    logger.error(f"Rasm yuborishda xatolik: {e}")
                    # Agar rasm yuborilmasa, oddiy matn yuborish
                    await query.edit_message_text(
                        f"💊 **{medicine['dori_nomi']}**\n"
                        f"💰 Нарх: {medicine['narxi']:,} сўм\n"
                        f"🏷️ ИКПУ: {medicine['ikpu']}\n\n"
                        f"ℹ️ **Маълумот:**\n{medicine.get('info', 'Маълумот мавжуд эмас')}\n\n"
                        "📝 Дори учун миқдор киритинг:",
                        parse_mode=ParseMode.MARKDOWN
                    )
            else:
                # Rasm mavjud emas, faqat matn
                await query.edit_message_text(
                    f"💊 **{medicine['dori_nomi']}**\n"
                    f"💰 Нарх: {medicine['narxi']:,} сўм\n"
                    f"🏷️ ИКПУ: {medicine['ikpu']}\n\n"
                    f"ℹ️ **Маълумот:**\n{medicine.get('info', 'Маълумот мавжуд эмас')}\n\n"
                    "📝 Дори учун миқдор киритинг:",
                    parse_mode=ParseMode.MARKDOWN
                )
        return

    # YANGI: Sahifalash
    if data.startswith("page_"):
        page = int(data.split("_")[1])
        medicines = db_manager.get_all_medicines()
        
        await query.edit_message_reply_markup(
            reply_markup=get_medicines_keyboard(medicines, page)
        )
        return

    # YANGI: Miqdorni tahrirlash
    if data.startswith("edit_quantity_"):
        medicine_id = int(data.split("_")[2])
        medicines = db_manager.get_all_medicines()
        medicine = next((m for m in medicines if m['id'] == medicine_id), None)
        
        if medicine:
            context.user_data['edit_medicine'] = medicine
            session.state = "edit_quantity"
            
            await query.edit_message_text(
                f"✏️ **{medicine['dori_nomi']}** дори учун янги миқдор киритинг:",
                parse_mode=ParseMode.MARKDOWN
            )
        return
    
    # YANGI: Orqaga qaytish
    if data == "back_to_medicines":
        medicines = db_manager.get_all_medicines()
        
        await query.edit_message_text(
            f"🏥 **{session.current_pharmacy['dorixona_nomi']}** учун буюртма\n\n"
            "💊 Қуйидаги дорилардан танланг:",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=get_medicines_keyboard(medicines)
        )
        return
    
    # YANGI: Buyurtmani yakunlash
    if data == "finish_order":
        if session.current_order:
            session.state = "payment_selection"
            order_summary = "📋 **Buyurtma xulosasi:**\n\n"
            for item in session.current_order:
                order_summary += f"💊 {item['name']} x {item['quantity']} = {item['total']:,} so'm\n"
            
            order_summary += f"\n💰 **Умумий сумма: {session.order_total:,} so'm**\n\n"
            order_summary += "Тўлов турини танланг:"
            
            await query.edit_message_text(
                order_summary,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_payment_keyboard()
            )
        else:
            await query.edit_message_text("❌ Буюртма бўш!")
        return
    
    # YANGI: Buyurtmani bekor qilish
    if data == "cancel_order":
        session.current_order = []
        session.order_total = 0
        session.state = "idle"
        
        await query.edit_message_text(
            "❌ Буюртма бекор қилинди!",
            reply_markup=None
        )
        return

    # To'lov tanlash (mavjud kod)
    if data.startswith("payment_"):
        if data == "payment_100":
            if session.order_total >= 6000000:
                session.discount_percentage = 8
            elif session.order_total >= 3000000:
                session.discount_percentage = 5
            else:
                session.discount_percentage = 0
        else:
            session.discount_percentage = 0
        
        discount_amount = session.order_total * session.discount_percentage / 100
        final_amount = session.order_total - discount_amount
        
        payment_info = f"💳 **Тўлов маълумотлари:**\n"
        payment_info += f"💰 Асосий сумма: {session.order_total:,} so'm\n"
        
        if session.discount_percentage > 0:
            payment_info += f"🎯 Чегирма: {session.discount_percentage}% ({discount_amount:,} so'm)\n"
            payment_info += f"💵 Якуний сумма: {final_amount:,} so'm\n"
        
        payment_info += f"📊 Тўлов: {'100%' if data == 'payment_100' else '50%'}\n\n"
        payment_info += "Буюртмани тасдиқлайсизми?"
        
        await query.edit_message_text(
            payment_info,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=get_confirm_keyboard()
        )

    # Tasdiqlash (mavjud kod)
    elif data.startswith("confirm_"):
        if data == "confirm_yes":
            # `context.application` ni `send_order_to_group` ga uzatish
            await send_order_to_group(context.application, session, user_id)

            session.current_order = []
            session.order_total = 0
            session.discount_percentage = 0
            session.state = "idle"
            
            await query.edit_message_text(
                "✅ Буюртма тасдиқланди ва гуруҳга юборилди!"
            )
        else:
            session.current_order = []
            session.order_total = 0
            session.discount_percentage = 0
            session.state = "idle"
            
            await query.edit_message_text(
                "❌ Буюртма бекор қилинди!"
            )




async def handle_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Geolokatsiya qayta ishlash"""
    user_id = update.effective_user.id
    
    if user_id not in user_sessions or not user_sessions[user_id].is_authorized:
        await update.message.reply_text("Аввал рўйхатдан ўтинг!")
        return
    
    session = user_sessions[user_id]
    location = update.message.location
    
    if not session.day_started:
        # Kun boshlash jarayoni
        session.location_received = True
        context.user_data['start_location'] = {
            'latitude': location.latitude,
            'longitude': location.longitude
        }
        
        await update.message.reply_text(
            "📍 Геолокация қабул қилинди!\n"
            "Энди камида 10 сония давомида видео хабар юборинг."
        )
    elif session.state == "waiting_pharmacy_location":
        # YANGI: Dorixona tasdiqlangandan keyingi geolokatsiya
        session.state = "idle"
        context.user_data['pharmacy_location'] = {
            'latitude': location.latitude,
            'longitude': location.longitude
        }
        
        await update.message.reply_text(
            "✅ Геолокация қабул қилинди!\n"
            "Энди асосий менюдан керакли бўлимни танлашингиз мумкин:",
            reply_markup=get_main_menu_keyboard()
        )
    else:
        # Kun tugashi jarayoni
        context.user_data['end_location'] = {
            'latitude': location.latitude,
            'longitude': location.longitude
        }
        
        await update.message.reply_text(
            "📍 Геолокация қабул қилинди!\n"
            "Энди камида 10 сония давомида видео хабар юборинг."
        )





async def send_order_to_group(application, session, user_id):
    """Buyurtmani guruhga spetsifikatsiya sifatida yuborish"""
    try:
        # Excel fayl yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Спецификация"

        # Sana
        ws.cell(row=1, column=2, value=f"Сана: {datetime.now().strftime('%d.%m.%Y')}").font = Font(bold=True)

        # Dorixona (pharmacy) ma'lumotlari
        pharmacy = session.current_pharmacy
        ws.cell(row=2, column=2, value=f"Приложение к дог №: {pharmacy['dagovor']}").font = Font(bold=True)

        # Поставщик
        supplier_data = [
            "ПОСТАВЩИК: MCHJ \"GYNOMEDIX\"",
            "АДРЕС: Toshkent shaxri Chilonzor tumani. Dumbirobod 4 tor kuchasi 23/2",
            "ТЕЛ.: 99 830-23-30",
            "ИНН: 311818897",
            "Р/с: 2020 8000 1071 8525 5001 МФО: 01095",
            "Регист. код плател. НДС: 326060260809"
        ]
        for idx, data in enumerate(supplier_data):
            ws.cell(row=3 + idx, column=1, value=data).font = Font(bold=True)

        # Покупатель
        mfo_value = pharmacy.get('mfo', 'N/A')
        mfo_formatted = str(int(float(mfo_value))).zfill(5) if mfo_value != 'N/A' and str(mfo_value).replace('.0', '').isdigit() else str(mfo_value)

        buyer_data = [
            f"ПОКУПАТЕЛЬ: {pharmacy['dorixona_nomi']}",
            f"АДРЕС: {pharmacy['manzil']}",
            f"ТЕЛ.: {pharmacy.get('telefon', 'N/A')}",
            f"ИНН: {pharmacy['inn']}",
            f"Р/с: {pharmacy.get('rs', 'N/A')}",
            f"банк мфо: {mfo_formatted}"
        ]
        for idx, data in enumerate(buyer_data):
            ws.cell(row=3 + idx, column=7, value=data).font = Font(bold=True)

        # Jadval
        table_start_row = 3 + max(len(supplier_data), len(buyer_data)) + 2
        headers = ["Товар", "ИКПУ", "Количество", "Цена", "Сумма без НДС", "НДС (12%)", "Скидка", "Итоговая сумма"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Ma'lumotlar
        current_row = table_start_row + 1
        total_original = total_nds = total_discount = total_final = 0

        for item in session.current_order:
            price_wo_nds = item['price'] / 1.12
            nds = item['price'] - price_wo_nds
            total_wo_nds = price_wo_nds * item['quantity']
            total_nds_item = nds * item['quantity']
            discount = item['total'] * session.discount_percentage / 100
            final = item['total'] - discount

            total_original += item['total']
            total_nds += total_nds_item
            total_discount += discount
            total_final += final

            ws.cell(row=current_row, column=1, value=item['name'])
            ws.cell(row=current_row, column=2, value=item['ikpu'])
            ws.cell(row=current_row, column=3, value=item['quantity'])
            ws.cell(row=current_row, column=4, value=round(item['price'], 2))
            ws.cell(row=current_row, column=5, value=round(total_wo_nds, 2))
            ws.cell(row=current_row, column=6, value=round(total_nds_item, 2))
            ws.cell(row=current_row, column=7, value=round(discount, 2))
            ws.cell(row=current_row, column=8, value=round(final, 2))
            current_row += 1

        # Итого
        ws.cell(row=current_row, column=4, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=current_row, column=5, value=round(total_original / 1.12, 2)).font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=round(total_nds, 2)).font = Font(bold=True)
        ws.cell(row=current_row, column=7, value=round(total_discount, 2)).font = Font(bold=True)
        ws.cell(row=current_row, column=8, value=round(total_final, 2)).font = Font(bold=True)

        # Imzolar
        signature_row = current_row + 3
        ws.cell(row=signature_row, column=1, value="ПОСТАВЩИК").font = Font(bold=True)
        ws.cell(row=signature_row + 1, column=1, value="Директор: RAXMONOV P.M. _______________").font = Font(bold=True)
        ws.cell(row=signature_row + 2, column=1, value="М.П").font = Font(bold=True)

        ws.cell(row=signature_row, column=7, value="ПОКУПАТЕЛЬ").font = Font(bold=True)
        ws.cell(row=signature_row + 1, column=7, value="Директор: ____________________").font = Font(bold=True)
        ws.cell(row=signature_row + 2, column=7, value="М.П").font = Font(bold=True)

        # Chegaralar
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(table_start_row, current_row + 1):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

        # Kengliklar
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['G'].width = 50
        for i, w in enumerate([30, 15, 12, 12, 15, 12, 12, 15], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Faylni saqlash
        filename = f"specification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        # Xodim ma’lumoti
        employee_name = session.employee_info.get('ism_familiya', 'Noma\'lum') if session.employee_info else 'Noma\'lum'

        # Caption
        caption = (
            f"📋 *Янги спецификация*\n"
            f"👤 Ходим: {employee_name}\n"
            f"🏥 Дорихона: {pharmacy['dorixona_nomi']}\n"
            f"💰 Асосий сумма: {total_original:,} so'm\n"
        )
        if session.discount_percentage > 0:
            caption += f"💸 Чегирма: {session.discount_percentage}% ({total_discount:,} so'm)\n"
        else:
            caption += f"💸 Чегирма: 0%\n"
        caption += f"💵 Якуний сумма: {total_final:,} so'm\n"
        caption += f"📅 Сана: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        # Guruhga yuborish (xatolikni tutish bilan)
        try:
            with open(filename, 'rb') as file:
                await application.bot.send_document(
                    chat_id=GROUP_CHAT_ID,
                    document=file,
                    filename=filename,
                    caption=caption,
                    parse_mode=ParseMode.MARKDOWN
                )
        except Exception as send_err:
            logger.error(f"Error sending specification to group: {send_err}")

        # Faylni o'chirish
        try:
            os.remove(filename)
        except Exception as remove_err:
            logger.warning(f"Could not delete file {filename}: {remove_err}")

    except Exception as e:
        logger.error(f"Unexpected error in send_order_to_group: {e}")


def main():
    """Asosiy funksiya"""
    # Bot ilovasini yaratish
    app = Application.builder().token(BOT_TOKEN).build()
    
    # Admin handlerlar
    app.add_handler(CommandHandler("admin", admin_panel))
    app.add_handler(CallbackQueryHandler(admin_callback_handler, pattern="^admin_"))
    
    # Asosiy handlerlar
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reset_auth", reset_auth))
    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    app.add_handler(MessageHandler(filters.LOCATION, handle_location))
    app.add_handler(MessageHandler(filters.VIDEO_NOTE, handle_video_note))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_message))
    app.add_handler(CallbackQueryHandler(handle_callback_query))
    app.bot_data['db_manager'] = db_manager
    
    # Botni ishga tushirish
    print("Bot ishga tushmoqda...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


    
if __name__ == '__main__':
    main()
